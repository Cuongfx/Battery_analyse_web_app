"""
Extract the HPPC pulse section from a Neware .xlsx export and write a
pipeline-compatible CSV for the cell ECM tool.

# Process a specific file (positional argument)
python extract_hppc.py path/to/MyCell.xlsx

# Pick a different sheet and output location
python extract_hppc.py path/to/MyCell.xlsx --sheet "Record List1" --output data/HPPC_data/MyCell.csv

# With plotting
python extract_hppc.py path/to/MyCell.xlsx --plot

# --input still works as an alternative
python extract_hppc.py --input path/to/MyCell.xlsx

# No argument falls back to the default Channel_8-3.xlsx
python extract_hppc.py

The Neware test sequence is assumed to be:
    full charge (CCC + CVC) -> rest -> HPPC pulses -> rest -> full charge.

The HPPC region is auto-detected as the block of repeated short current
pulses between the initial and final full-charge phases. Each SOC section in
the region follows the motif:
    discharge pulse -> rest -> (charge pulse -> rest) -> constant discharge -> long rest

The output CSV uses the columns expected by `ecm/data/hppc.py`:
    Time(s), Current(A), Voltage(V), Data
where the `Data` column contains the `S` markers (5 per SOC section) that the
fitting pipeline uses to locate pulse and discharge sections.
"""

import argparse
from pathlib import Path

import numpy as np

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to read .xlsx files. Install it with "
        "`pip install openpyxl` or `pip install -r requirements.txt`."
    ) from exc


# Column names in the Neware "Record List" sheet.
COL_COMMAND = "Command"
COL_TYPE = "Cmd Type"
COL_TIME = "Record Time"
COL_VOLTAGE_MV = "V(mV)"
COL_CURRENT_A = "I(A)"

# Default classification thresholds.
DEFAULT_PULSE_MAX_SECONDS = 60.0    # steps shorter than this are pulses
FULL_CHARGE_MIN_SECONDS = 600.0     # CCC longer than this is a full charge
DEFAULT_DT = 0.1                    # fallback sample period [s]


def parse_record_time(value):
    """
    Parse a Neware `Record Time` string (`H:MM:SS.fff`) into seconds.
    """
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    parts = text.split(":")
    try:
        if len(parts) == 3:
            hours, minutes, seconds = parts
            return int(hours) * 3600 + int(minutes) * 60 + float(seconds)
        if len(parts) == 2:
            minutes, seconds = parts
            return int(minutes) * 60 + float(seconds)
        return float(parts[0])
    except ValueError:
        return None


def load_records(path, sheet_name):
    """
    Stream the requested sheet and return cleaned per-row arrays.

    Returns a dict with numpy arrays: command, cmd_type, time_s (continuous),
    current, voltage. Blank separator rows are dropped.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}
    for required in (COL_COMMAND, COL_TYPE, COL_TIME, COL_VOLTAGE_MV, COL_CURRENT_A):
        if required not in idx:
            raise SystemExit(
                f"Column '{required}' not found in sheet. Found columns: {list(header)}"
            )

    commands = []
    cmd_types = []
    record_times = []
    currents = []
    voltages = []

    for row in rows:
        command = row[idx[COL_COMMAND]]
        cmd_type = row[idx[COL_TYPE]]
        # Drop blank separator rows inserted between steps.
        if command is None or cmd_type is None:
            continue
        rt = parse_record_time(row[idx[COL_TIME]])
        current = row[idx[COL_CURRENT_A]]
        voltage_mv = row[idx[COL_VOLTAGE_MV]]
        if rt is None or current is None or voltage_mv is None:
            continue
        commands.append(float(command))
        cmd_types.append(str(cmd_type))
        record_times.append(rt)
        currents.append(float(current))
        voltages.append(float(voltage_mv) / 1000.0)

    wb.close()

    if not commands:
        raise SystemExit("No usable data rows found in sheet.")

    commands = np.array(commands)
    record_times = np.array(record_times)
    time_s = _continuous_time(commands, record_times)

    return {
        "command": commands,
        "cmd_type": np.array(cmd_types, dtype=object),
        "time_s": time_s,
        "current": np.array(currents),
        "voltage": np.array(voltages),
    }


def _continuous_time(commands, record_times):
    """
    Build a monotonic global time vector from per-step `Record Time` values,
    which reset to zero at the start of every step.
    """
    n = len(record_times)
    time_s = np.zeros(n)
    last_dt = DEFAULT_DT
    for k in range(1, n):
        same_step = commands[k] == commands[k - 1]
        delta = record_times[k] - record_times[k - 1]
        if same_step and delta > 0:
            last_dt = delta
            step = delta
        else:
            step = last_dt
        time_s[k] = time_s[k - 1] + step
    return time_s


def build_blocks(commands):
    """
    Group consecutive rows that share the same Command value.

    Returns a list of (start_index, end_index_inclusive) tuples.
    """
    blocks = []
    start = 0
    for k in range(1, len(commands)):
        if commands[k] != commands[start]:
            blocks.append((start, k - 1))
            start = k
    blocks.append((start, len(commands) - 1))
    return blocks


def classify_block(data, block, pulse_max_seconds):
    """
    Classify a block as one of: full_charge, charge_pulse, discharge_pulse,
    constant_discharge, rest, other.
    """
    start, end = block
    cmd_type = str(data["cmd_type"][start]).upper()
    duration = data["time_s"][end] - data["time_s"][start]
    current = data["current"][start:end + 1]
    mean_current = float(np.mean(current))

    if cmd_type == "REST" or np.allclose(current, 0.0):
        return "rest"

    # Charge steps have positive current (CCC/CVC).
    if mean_current > 0 or cmd_type in ("CCC", "CVC"):
        if duration >= FULL_CHARGE_MIN_SECONDS or cmd_type == "CVC":
            return "full_charge"
        return "charge_pulse"

    # Discharge steps have negative current (CCD).
    if duration <= pulse_max_seconds:
        return "discharge_pulse"
    return "constant_discharge"


def detect_hppc_region(data, blocks, pulse_max_seconds):
    """
    Find the row range covering the HPPC pulses.

    Returns (end_row_exclusive, section_flags) where section_flags is a list of
    (flag1..flag5) row indices per SOC section, matching the layout expected by
    `ecm/data/hppc.py`. Each "start" flag is placed on the last rested sample
    just before the step so the loader sees a real current step (for R0) and a
    rested OCV anchor.
    """
    kinds = [classify_block(data, b, pulse_max_seconds) for b in blocks]

    # First discharge pulse marks the start of the HPPC region.
    first_pulse = next((i for i, k in enumerate(kinds) if k == "discharge_pulse"), None)
    if first_pulse is None:
        raise SystemExit("Could not find any discharge pulse to start the HPPC region.")
    if first_pulse == 0:
        raise SystemExit("HPPC region starts at the first block; no preceding rest found.")

    # Final full charge marks the end of the HPPC region.
    final_charge = None
    for i in range(len(kinds) - 1, first_pulse, -1):
        if kinds[i] == "full_charge":
            final_charge = i
        elif final_charge is not None and kinds[i] not in ("full_charge", "rest"):
            break
    if final_charge is None:
        end_row = blocks[-1][1] + 1
        region_block_end = len(blocks)
    else:
        end_row = blocks[final_charge][0]
        region_block_end = final_charge

    # Walk blocks within the region and build 5 flags per SOC section.
    section_flags = []
    i = first_pulse
    while i < region_block_end:
        if kinds[i] != "discharge_pulse":
            i += 1
            continue

        d_start, d_end = blocks[i]

        # flag1: last rested sample before the discharge pulse (rested OCV).
        flag1 = blocks[i - 1][1] if kinds[i - 1] == "rest" else d_start
        # flag2: end of the discharge pulse.
        flag2 = d_end
        # flag3: end of the rest after the discharge pulse.
        flag3 = blocks[i + 1][1] if (i + 1 < region_block_end and kinds[i + 1] == "rest") else d_end

        # Find the constant discharge that ends this SOC section.
        const_idx = None
        j = i + 1
        while j < region_block_end:
            if kinds[j] == "discharge_pulse":
                break  # next section started; no constant discharge found
            if kinds[j] == "constant_discharge":
                const_idx = j
                break
            j += 1

        if const_idx is None:
            i = j
            continue

        c_start, c_end = blocks[const_idx]
        # flag4: last rested sample before the constant discharge (for R0 step).
        flag4 = blocks[const_idx - 1][1] if kinds[const_idx - 1] == "rest" else c_start - 1
        # flag5: end of the constant discharge (start of the relaxation tail).
        flag5 = c_end
        section_flags.append((flag1, flag2, flag3, flag4, flag5))
        i = const_idx + 1

    if not section_flags:
        raise SystemExit("No complete SOC sections (pulse + constant discharge) found.")

    return end_row, section_flags


def build_output(data, end_row, section_flags):
    """
    Build the output arrays (time zeroed at region start) and the `Data` flag
    column. One extra rested sample is prepended as the leading global marker
    that `CellHppcData.__init__` drops.
    """
    flag_rows = sorted({flag for section in section_flags for flag in section})
    # Drop the final constant-discharge marker so the flag layout matches the
    # loader's expectation (last SOC section has a pulse but no fitted
    # discharge), keeping pulse/discharge index counts aligned.
    flag_rows = flag_rows[:-1]

    first_flag = flag_rows[0]
    lead_row = max(first_flag - 1, 0)

    rows = np.arange(lead_row, end_row)
    time = data["time_s"][rows]
    time = time - time[0]
    current = data["current"][rows]
    voltage = data["voltage"][rows]

    flags = np.full(len(rows), "", dtype=object)
    flags[0] = "S"  # leading global marker (dropped by the loader)

    row_to_local = {row: i for i, row in enumerate(rows)}
    for flag_row in flag_rows:
        local = row_to_local.get(flag_row)
        if local is not None:
            flags[local] = "S"

    return time, current, voltage, flags


def build_full_output(data, section_flags):
    """
    Build whole-process output spanning the entire test (initial full charge ->
    rest -> HPPC -> final rest -> recharge), so the captured CSV/plot show the
    complete process rather than only the cropped HPPC window.

    The HPPC section markers use the same layout as `build_output`, so per-SOC
    pulse fitting is unchanged. The leading and trailing full-charge regions
    carry no `S` markers: the loader (`CellHppcData`) keeps from the second
    marker onward, so the leading charge is ignored automatically, and the fit
    wrapper clips the trailing charge at the last marker.
    """
    flag_rows = sorted({flag for section in section_flags for flag in section})
    # Drop the final constant-discharge marker for the same count-alignment
    # reason as build_output (last SOC section has a pulse but no fitted discharge).
    flag_rows = flag_rows[:-1]

    n = len(data["time_s"])
    time = data["time_s"][:n] - data["time_s"][0]
    current = data["current"][:n]
    voltage = data["voltage"][:n]

    flags = np.full(n, "", dtype=object)
    flags[0] = "S"  # leading global marker (dropped by the loader)
    for flag_row in flag_rows:
        if 0 <= flag_row < n:
            flags[flag_row] = "S"

    return time, current, voltage, flags


def write_csv(path, time, current, voltage, flags):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as fh:
        fh.write("Time(s),Current(A),Voltage(V),Data\n")
        for t, i, v, flag in zip(time, current, voltage, flags):
            fh.write(f"{t:.3f},{i:.6f},{v:.6f},{flag}\n")


def plot_region(time, current, voltage, save_path=None, show=True, title="Extracted HPPC pulses"):
    import matplotlib.pyplot as plt

    fig, (ax_i, ax_v) = plt.subplots(2, 1, figsize=(10, 6), sharex=True, tight_layout=True)
    ax_i.plot(time, current, color="C0")
    ax_i.set_ylabel("Current [A]")
    ax_i.set_title(title)
    ax_i.grid(True, color="0.9")

    ax_v.plot(time, voltage, color="C3")
    ax_v.set_xlabel("Time [s]")
    ax_v.set_ylabel("Voltage [V]")
    ax_v.grid(True, color="0.9")

    if save_path is not None:
        save_path = Path(save_path)
        save_path.parent.mkdir(parents=True, exist_ok=True)
        fig.savefig(save_path, dpi=150)

    if show:
        plt.show()

    return fig


def build_parser():
    parser = argparse.ArgumentParser(
        description="Extract HPPC pulses from a Neware .xlsx export into a pipeline-ready CSV."
    )
    parser.add_argument(
        "input",
        type=Path,
        nargs="?",
        default=Path("Channel_8-3.xlsx"),
        help="Path to the Neware .xlsx file to process.",
    )
    parser.add_argument(
        "--input",
        type=Path,
        dest="input_opt",
        help="Alternative way to specify the .xlsx file (overrides the positional argument).",
    )
    parser.add_argument(
        "--sheet",
        default="Record List1",
        help="Worksheet name containing the record list.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output CSV path. Defaults to data/HPPC_data/<input-name>.csv.",
    )
    parser.add_argument(
        "--pulse-max-seconds",
        type=float,
        default=DEFAULT_PULSE_MAX_SECONDS,
        help="Steps shorter than this (seconds) are treated as pulses.",
    )
    parser.add_argument(
        "--plot",
        action="store_true",
        help="Plot the extracted pulse current and voltage curves.",
    )
    parser.add_argument(
        "--save-plot",
        type=Path,
        help="Optional image path to save the extracted pulse plot.",
    )
    parser.add_argument(
        "--no-show",
        action="store_true",
        help="Do not open a plot window (useful with --save-plot).",
    )
    return parser


def main():
    args = build_parser().parse_args()

    input_path = args.input_opt or args.input
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")

    output = args.output or Path("data/HPPC_data") / (input_path.stem + ".csv")

    print(f"Reading: {input_path} [sheet: {args.sheet}]")
    data = load_records(input_path, args.sheet)
    print(f"Loaded {len(data['command'])} data rows (blank rows dropped).")

    blocks = build_blocks(data["command"])
    end_row, section_flags = detect_hppc_region(
        data, blocks, args.pulse_max_seconds
    )
    print(f"Detected {len(section_flags)} SOC sections; region ends at row {end_row - 1}.")

    time, current, voltage, flags = build_output(data, end_row, section_flags)
    write_csv(output, time, current, voltage, flags)
    print(f"Wrote {len(time)} rows to: {output}")
    print(f"Duration: {time[-1]:.1f} s, voltage {voltage.min():.3f}-{voltage.max():.3f} V, "
          f"current {current.min():.2f}-{current.max():.2f} A")

    if args.plot or args.save_plot:
        plot_region(
            time,
            current,
            voltage,
            save_path=args.save_plot,
            show=not args.no_show,
        )
        if args.save_plot:
            print(f"Saved pulse plot: {args.save_plot}")


if __name__ == "__main__":
    main()
